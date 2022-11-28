from openpyxl import Workbook, load_workbook
import pandas as pd


class ExcelWriter():
    def __init__(self, path):
        self.counts = None
        self.path = path
        self.__create_file(path)

    def __create_file(self, path):
        wb = Workbook()
        wb.save(self.path)

    def __convert_to_pd(self):
        log_df = pd.DataFrame(self.counts[0])
        db_df = pd.DataFrame(self.counts[1])
        return pd.concat([log_df, db_df], axis=1)

    def __add_header(self, writer, sheetname, columns):
        for col_num, value in enumerate(columns):
            writer.sheets[sheetname].cell(
                row=1, column=col_num + 1).value = value

    def __load_excel(self):
        wb = load_workbook(self.path)
        ws = wb.active
        return wb, ws

    def write_to_excel(self, counts, write=True):
        self.counts = counts
        df = self.__convert_to_pd()
        columns = ['Nazwa pliku', 'Klasa', 'Count_LOGI', 'Count_BAZA']

        if write:
            wb, ws = self.__load_excel()

            with pd.ExcelWriter(self.path, engine='openpyxl') as writer:
                writer.book = wb
                writer.sheets = {ws.title: ws for ws in wb.worksheets}

                for sheetname in writer.sheets:
                    df.to_excel(writer, sheet_name=sheetname,
                                startrow=writer.sheets[sheetname].max_row, index=False, header=False)

                self.__add_header(writer, sheetname, columns)

                writer.book.save(self.path)

    def createDataFrame(self, fileList, dataBaseList):
        data = [fileList, dataBaseList]
        dataFrame = pd.DataFrame(data).T
        return dataFrame

    def concatDataFrames(self, filesDF, comparedfilesDF):
        concatData = pd.concat([filesDF, comparedfilesDF], axis=1)
        return concatData

    def write_another_sheet(self, filesDF, comparedfilesDF, invalidList, sheetname, write=True):
        if sheetname == 'Files_with_LOGS':
            columns = ['Files LOGS', 'Files DataBase',
                       'Difference Files-DataBase', 'Difference DataBase-Files', 'Invalid Files']

        if sheetname == 'Files_IN_Folder':
            columns = ['Files', 'Files DataBase',
                       'Difference Files-DataBase', 'Difference DataBase-Files', 'Invalid Files']

        data = [invalidList]
        invalidDf = pd.DataFrame(data).T

        dataFramefiles = self.concatDataFrames(filesDF, comparedfilesDF)
        dataFrame = pd.concat([dataFramefiles, invalidDf], axis=1)
        dataFrame.columns = columns

        with pd.ExcelWriter(self.path, engine='openpyxl', mode='a') as writer:
            dataFrame = dataFrame.sort_values(by=columns[0], ascending=True)
            dataFrame.to_excel(writer, sheet_name=sheetname,
                               startrow=1, index=False, header=False)
            self.__add_header(writer, sheetname, columns)
